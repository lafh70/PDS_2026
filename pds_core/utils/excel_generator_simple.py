"""Simple Excel template generator for Phase 1 ingestion."""

from pathlib import Path
from typing import Dict, List, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pds_core.config.loader_columnas_simple import CargadorColumnasSimple
from pds_core.config.schema.columnas_default import ColumnaDef


class GeneradorExcelIngestaSimple:
    """Generates a multi-sheet Excel ingestion template with openpyxl."""

    INSTRUCCIONES = [
        "PDS 2026 — Plantilla de ingesta",
        "",
        "1. Complete las pestañas GESTION, SOURCING, PERMITS y FINANCE.",
        "2. Project_ID es obligatorio en todas las pestañas (formato: HK-9001-2026-01).",
        "3. Columnas marcadas con * son obligatorias.",
        "4. Guarde el archivo y ejecute el script de ingesta.",
        "",
        "Pestañas:",
        "  • GESTION  — Datos maestros del proyecto (H6, H11 reales)",
        "  • SOURCING — Licitaciones DDO y GC",
        "  • PERMITS  — Trámites municipales (H8, H9, H10)",
        "  • FINANCE  — Presupuesto y contabilidad",
    ]

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def __init__(self, client_name: str = "hackathon"):
        self.client_name = client_name

    @staticmethod
    def generate(
        columnas: Dict[str, List[ColumnaDef]],
        output_path: Union[str, Path],
    ) -> Path:
        """
        Create an openpyxl workbook with INSTRUCCIONES + data sheets.

        Args:
            columnas: Dict mapping sheet name to list of ColumnaDef.
            output_path: Destination .xlsx path.

        Returns:
            Path to the written file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws_instr = wb.active
        ws_instr.title = "INSTRUCCIONES"

        for row_idx, linea in enumerate(GeneradorExcelIngestaSimple.INSTRUCCIONES, start=1):
            ws_instr.cell(row=row_idx, column=1, value=linea)
        ws_instr.column_dimensions["A"].width = 80

        for tabla in ("GESTION", "SOURCING", "PERMITS", "FINANCE"):
            cols = columnas.get(tabla, [])
            GeneradorExcelIngestaSimple._crear_hoja_datos(wb, tabla, cols)

        wb.save(output_path)
        return output_path

    def generar(self, output_path: Union[str, Path]) -> Path:
        """Load columns for this client and generate the template."""
        columnas = CargadorColumnasSimple.load(self.client_name)
        return self.generate(columnas, output_path)

    @staticmethod
    def _crear_hoja_datos(
        wb: Workbook, tabla: str, columnas: List[ColumnaDef]
    ) -> None:
        ws = wb.create_sheet(title=tabla)
        sorted_cols = sorted(columnas, key=lambda c: c.orden_default)

        for col_idx, col_def in enumerate(sorted_cols, start=1):
            header = col_def.nombre_display or col_def.nombre_canónico
            if col_def.obligatorio:
                header = f"* {header}"

            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = GeneradorExcelIngestaSimple.HEADER_FILL
            cell.font = GeneradorExcelIngestaSimple.HEADER_FONT

            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = max(len(header) + 2, 14)

            if col_def.validación_tipo == "dropdown" and col_def.valores_desde:
                valores = ",".join(col_def.valores_desde)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{valores}"',
                    allow_blank=not col_def.obligatorio,
                )
                dv.add(f"{letter}2:{letter}1000")
                ws.add_data_validation(dv)

            if col_def.descripción:
                from openpyxl.comments import Comment
                cell.comment = Comment(col_def.descripción, "PDS")
