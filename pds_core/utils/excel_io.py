"""Atomic Excel read/write preserving other sheets."""

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


def read_sheet(filepath: str | Path, sheet_name: str, header_row: int = 1) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {filepath}")
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row - 1, engine="openpyxl")
        df.columns = df.columns.astype(str).str.strip()
        return df
    except PermissionError:
        raise PermissionError(f"File is open/locked: {filepath}. Close it and retry.") from None


def write_sheet(filepath: str | Path, sheet_name: str, df: pd.DataFrame) -> None:
    path = Path(filepath)
    tmp_path = path.with_suffix(".tmp.xlsx")
    ensure_output_dir(path)

    if path.exists():
        try:
            wb = load_workbook(path)
        except PermissionError:
            raise PermissionError(f"File is open/locked: {filepath}. Close it and retry.") from None
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    _df_to_ws(df, ws)
    try:
        wb.save(tmp_path)
    finally:
        wb.close()
    shutil.move(str(tmp_path), str(path))


def read_workbook(filepath: str | Path) -> dict[str, pd.DataFrame]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")
    xl = pd.ExcelFile(path, engine="openpyxl")
    return {sheet: pd.read_excel(path, sheet_name=sheet, engine="openpyxl") for sheet in xl.sheet_names}


def ensure_output_dir(filepath: str | Path) -> None:
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)


def _df_to_ws(df: pd.DataFrame, ws) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
